#!/usr/bin/python

'''
/*
 * @File: utils.py
 * @Author: Arunkumar Sadasivan
 * @Date: 11/01/2017
 * @Description: utils.py contains common methods that are used by different python modules and classes.
 * @Usage: N/A
 */
 
'''
import os
import sys
import datetime
import configparser
import json
import re
import importlib
# python3 -m pip install openpyxl
import openpyxl 

# Defaults
csvdelimiter = "@@" # data might contain commas, semicolon etc. So, it is safe to use a delimiter that doesn't exists in the string.
eoldelimiter = "@@@" # end of line delimiter

'''
########################## Config Parser #################################################################################
'''
# Read configuration file
def readConfig(configFile):
    """
    Returns the config object
    """
    if not os.path.exists(configFile):
        print("[Error] Unable to find the configuration file:" + configFile)
        exit
    else:
        config = configparser.ConfigParser()
        config.read(configFile)
        return config   # returns config object
    

# get the config value
def getConfigValue(configObj,section,key): 
    return configObj.get(section,key)


'''
########################## JSON Parser #################################################################################
'''

# Read JSON File
def readJSONFile(file):
    with open(file, 'r') as json_file:
        json_data = json.load(json_file)
        return json_data

# Create JSON file   
def writeJSONFile(file, json_data):
    with open(file, 'w') as outfile:
        json.dump(json_data, outfile, sort_keys=False, indent=4)  
    print("[Done] Guidelines automation configuration file successfully created: " + file)      

# This method is used to print the JSON data in a much more human readable format    
def prettyPrintJSON(json_data):
    print(json.dumps(json_data, sort_keys=True, indent=4))


'''
########################## Excel Parser #################################################################################
'''
# read the sheet and convert it into CSV
def getGuidelinesArry(guidelineFileName, guideLineName):
    workbook = openpyxl.load_workbook(guidelineFileName)
    sheetName = getSheetName(workbook)
    sheet = workbook[sheetName]
#   rows = sheet.rows 
#   rows = sheet.max_row
#   column = sheet.max_column
    '''
    row[0] = Section Name, row[1] = Section, row[2] = Guideline, 
    row[3] = Guideline Values, row[4] = Validation, row[5] = Key
    '''
    guidelineCSV = []
    for row in sheet.iter_rows(row_offset=15): # skip rows including header 
        guideLine = row[2].value
        if guideLine:
            guideLine = guideLine.strip()
            sectionName = "Section Name: " + row[0].value 
            section = "Section: " + str(row[1].value) 
            automationKey = row[4].value
            guidelineCSV.append([sectionName, section, guideLine, automationKey, guideLineName])
    return guidelineCSV

# get sheetname
def getSheetName(workbook):
    # get sheet names array
    sheetnames = workbook.sheetnames
    for sheetname in sheetnames:
        if sheetname == "revision": #skip 
            continue
        else:
            return sheetname

'''
########################## File Parser #################################################################################
'''
# Create directory if it doesn't exist
def checkDirExists(path):
    if not os.path.exists(path): # check directory exists 
        os.makedirs(path) # create directory

'''
########################## Calling function of a module using the module name and function name as string ###############
'''
'''
Note:
* function paarameters need to be a list.
* Using the *<varaibaleName> to call with arguments unpacked from the list. 
* Check this link for more details: 
  https://docs.python.org/3/tutorial/controlflow.html#unpacking-argument-lists
'''
def methodCallUsingString(moduleName, functionName, parameterList):
    '''
    method_to_call = getattr(moduleName, functionName)
    if callable(method_to_call):
        result = method_to_call()
    '''    
    methodCall = None
    try:
        moduleObj = importlib.import_module(moduleName) # function returns the package object
        methodCall = getattr(moduleObj, functionName)(*parameterList) # Unpacking Argument Lists
    except ImportError as err:
        print("[Error] Module missing:" + moduleName) 
        print(err)
        sys.exit() 
    except AttributeError as err:
        print("[Error] Unable to find the function:"+ functionName + ". Make sure function exists in " + moduleName)
        print(err)
        sys.exit() 
        
    return methodCall

    
'''
########################## Date/Time #################################################################################
'''
   
# This method is to get the current date and time
def getDateTime(fileName,fileExt):
    # get current date and time 
    currentDateTime = datetime.datetime.now()
    # append the file with current date and time
    fileName = fileName + "_" + currentDateTime.strftime("%Y-%m-%d_%H-%M-%S") + "." + fileExt
    return fileName

'''
########################## HTML Report #################################################################################
'''

# Convert a CSV data into html report. 
def createHTMLReport(csvdata, reportFile, reportTitle, detailsArry):
    csvArry = csvdata.split(eoldelimiter) # convert the CSV string into list)
    csvArry = filter(None, csvArry) # remove empty values
    print("[Info] Creating HTML report...")
    with open(reportFile, 'w') as fileHandle: #enter the output filename
        fileHandle.write("<head>" + "\n")
        fileHandle.write("<!-- The line below is to make sure it uses the right encoding format  -->" + "\n")
        fileHandle.write("<meta http-equiv='Content-Type' content='application/xhtml+xml; charset=UTF-8' />" + "\n")
        fileHandle.write("</head>" + "\n")
        
        fileHandle.write("<!-- Latest compiled and minified CSS -->" + "\n")
        fileHandle.write("<link rel='stylesheet' href='https://cdnjs.cloudflare.com/ajax/libs/bootstrap-table/1.8.1/bootstrap-table.min.css'>" + "\n") 
        fileHandle.write("<!-- Latest compiled and minified CSS -->" + "\n")
        fileHandle.write("<link rel='stylesheet' href='https://maxcdn.bootstrapcdn.com/bootstrap/3.3.5/css/bootstrap.min.css'>" + "\n")       
       
        fileHandle.write("<table data-toggle = 'table' data-pagination = 'true'>" + "\n")
        fileHandle.write("<h3 align='center'> <font color='blue'> " + reportTitle +  " </font> </h3>" + "\n")
        if detailsArry:
            for detail in detailsArry:
                fileHandle.write("<h4 align='left'> <font color='black'> " + detail +  " </font> </h4>" + "\n")
        fileHandle.write("<style type='text/css'>" + "\n")
        fileHandle.write("table { color: #333; /* Lighten up font color */ font-family: Helvetica, Arial, sans-serif; /* Nicer font */  width: 100%; border-collapse:collapse; border-spacing: 2; }"+ "\n")
        fileHandle.write("table, td, th, tr { border-style: solid; border-width: 2px;}" + "\n")
        fileHandle.write("td, th { height: 30px;} /* Make cells a bit taller */" + "\n")
        fileHandle.write("th { background: #FFFFFF; /* Light grey background */ font-weight: bold; /* Make sure they\"re bold */ text-align: left; /* align text to left */} " + "\n")
        fileHandle.write("td { background: #DFDFDF; /* White background */ text-align: left; /* align text to left */ } " + "\n")
        fileHandle.write("</style>" + "\n")
        firstrow = True
        for row in csvArry:
            rows = row.split(csvdelimiter)
            if firstrow:
                fileHandle.write("\n\n\n\n\t" + "<thead>" + "\n" + "\t\t" + "<tr>" + "\n")
                for col in rows:
                    fileHandle.write("\t\t\t" + "<th data-sortable='true'>" +  col + "</th>" + "\n")
                fileHandle.write("\t\t" + "</tr>" + "\n\t" + "</thead>"  + "\n")
                fileHandle.write("\t" + "<tbody>" + "\n")
                firstrow = False
            else:
                fileHandle.write("\t\t" + "<tr>" + "\n")
                for col in rows:
                    col = col.replace("\n","<br />")
                    fileHandle.write("\t\t\t" + "<td>"  + col + "</td>" + "\n")
                fileHandle.write("\t\t" + "</tr>" + "\n")
        fileHandle.write("\t" + "</tbody>" + "\n")
        fileHandle.write("</table>" + "\n")
        
        fileHandle.write("<!-- jQuery (necessary for Bootstrap\'s JavaScript plugins) -->" + "\n")
        fileHandle.write("<script src='https://ajax.googleapis.com/ajax/libs/jquery/2.1.0/jquery.min.js'></script>" + "\n")
        fileHandle.write("<!-- Latest compiled and minified JavaScript -->"  + "\n")
        fileHandle.write("<script src='https://maxcdn.bootstrapcdn.com/bootstrap/3.3.5/js/bootstrap.min.js'></script>" + "\n")
        fileHandle.write("<!-- Latest compiled and minified JavaScript -->" + "\n")
        fileHandle.write("<script src='https://cdnjs.cloudflare.com/ajax/libs/bootstrap-table/1.8.1/bootstrap-table.min.js'></script>" + "\n")
    
    # Docker adds a "/" character when bind mount is performed. Removing the char "/" if exists
    if reportFile[0] == "/":
        reportFile = reportFile[1:] # remove first character alone
    print("[Done] Custom HTML report successfully created: " + reportFile)
    

'''
########################## CSV Report #################################################################################
'''    
# Convert a CSV data into html report. 
def createCSVReport(csvdata, reportFile, detailsArry):
#     csvArry = csvdata.split(eoldelimiter) # convert the CSV string into list)
#     csvArry = filter(None, csvArry) # remove empty values
    print("[Info] Creating CSV report...")
    jsonArry = []
    outputData = {}
    detailsDict = convertDetailstoDict(detailsArry)
    violationsArry = convertCSVDataintoDict(csvdata)
    outputData["violations"] = violationsArry
    outputData["details"] = detailsDict
    jsonArry.append(outputData)
    with open(reportFile, 'w', encoding = "utf-8") as fileHandle: #enter the output filename
        json.dump(jsonArry, fileHandle, indent = 4, sort_keys=True)       
     # Docker adds a "/" character when bind mount is performed. Removing the char "/" if exists
    if reportFile[0] == "/":
        reportFile = reportFile[1:] # remove first character alone                
    print("[Done] Custom CSV report successfully created: " + reportFile)   
     
    
# convert details Array to dictionary
def convertDetailstoDict(detailsArry):
    detailsDict = {}
    regexPattern = "(.*?):(.*?)$"
    for detail in detailsArry:
        matchObj = re.match(regexPattern, detail)
        if matchObj:
            detailKey = matchObj.group(1)
            detailValue = matchObj.group(2)
            detailsDict[detailKey] = detailValue.strip()
    return detailsDict   

# convert csv data into dictionary
def convertCSVDataintoDict(csvdata):
    csvArry = csvdata.split(eoldelimiter) # convert the CSV string into list)
    csvArry = filter(None, csvArry) # remove empty values
    firstrow = True 
    violationsArry = []
    headerKey = None
    for row in csvArry:
        rows = row.split(csvdelimiter)  
        #rowSize = len(rows)
        incrementer = 0
        if firstrow: # header
            firstrow = False
            headerKey = rows
            continue
        else:
            violationsDict = {}
            for col in rows:
                violationsDict[headerKey[incrementer]] = col
                incrementer = incrementer + 1
            violationsArry.append(violationsDict)       
    return  violationsArry
